import os
import re
import pandas as pd
import matplotlib.pyplot as plt
import textwrap
from openpyxl import load_workbook

# Function to install required packages
def install_packages():
    import subprocess
    import sys
    subprocess.check_call([sys.executable, "-m", "pip", "install", "pandas", "openpyxl", "matplotlib"])

try:
    import pandas as pd
    import matplotlib.pyplot as plt
    from openpyxl import load_workbook
except ImportError:
    install_packages()

def replace_keywords(text):
    replacements = {
        "zonestante": "ZS",
        "airesante": "AS",
        "localite": "LOCALITE",
        "_": " "
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    return text

def wrap_text(text, width):
    return '\n'.join(textwrap.wrap(text, width))


def process_excel_files(directory, max_files=None):
    count = 0

    output_dir = os.path.join(directory, 'table_img')
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    for filename in os.listdir(directory):
        if "_km" in filename.lower() and filename.endswith('.xlsx'):
            if max_files is not None and count >= max_files:
                break
            filepath = os.path.join(directory, filename)
            # Process the filename to create a valid output image name
            base_name = filename.split('_')[0]
            safe_name = re.sub(r"[ ']", lambda x: '-' if x.group() == ' ' else '', base_name)
            output_filename = f"{safe_name}.jpg"

            # Read the Excel file
            wb = load_workbook(filepath, data_only=True)
            ws = wb.active

            # Collect data from the sheet
            data = []
            headers = []
            is_header = True  # Flag to identify the header row

            for row in ws.iter_rows(values_only=True):
                if any(cell is not None for cell in row):
                    if is_header:
                        headers = [replace_keywords(str(cell)) if cell is not None else '' for cell in row]  # Set header
                        is_header = False  # Update flag so only the first row is treated as header
                    else:
                        data.append([replace_keywords(wrap_text(str(cell), 20)) if cell is not None else '' for cell in row])

            # Convert data to DataFrame with headers set properly
            df = pd.DataFrame(data, columns=headers)

            # Plotting the DataFrame
            fig, ax = plt.subplots(figsize=(8, 6))  # Adjust figsize dynamically based on data if needed
            ax.axis('tight')
            ax.axis('off')
            table = ax.table(cellText=df.values, colLabels=df.columns, cellLoc='center', loc='center')

            # Apply alternating row backgrounds and bold font for header
            for (i, j), cell in table.get_celld().items():
                if i == 0:
                    cell.set_facecolor("#f2f2f2")
                    cell.get_text().set_weight('bold')
                    cell.get_text().set_fontsize(12)
                    cell.get_text().set_family('Arial')
                elif i % 2 == 0:
                    cell.set_facecolor("#f9f9f9")  # Light grey for even rows
                else:
                    cell.set_facecolor("w")  # White for odd rows

            output_filename = os.path.join(output_dir, f"{safe_name}.jpg")
            plt.savefig(output_filename, dpi=300) 
            plt.close()

            print(f"Saved image as {output_filename}")
            count += 1

# params
directory = "E:\mheaton\cartography\COD_microplanning_042024\data\input\distance_matrix_by_AS"
process_excel_files(directory, max_files=5)
